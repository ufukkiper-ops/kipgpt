import base64
import csv
import io
import os

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

FILE_ICONS = {
    "pdf": "📄",
    "image": "🖼️",
    "word": "📝",
    "excel": "📊",
    "text": "📃",
    "other": "📎",
}


def get_file_category(filename, mimetype=""):
    ext = os.path.splitext(filename or "")[1].lower()
    if ext in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp") or mimetype.startswith("image/"):
        return "image"
    if ext == ".pdf" or mimetype == "application/pdf":
        return "pdf"
    if ext in (".doc", ".docx") or "word" in mimetype:
        return "word"
    if ext in (".xls", ".xlsx", ".csv") or "spreadsheet" in mimetype or "excel" in mimetype:
        return "excel"
    if ext in (".txt", ".md", ".log"):
        return "text"
    return "other"


def file_icon_for(category):
    return FILE_ICONS.get(category, FILE_ICONS["other"])


def _read_file_bytes(file_storage):
    if hasattr(file_storage, "seek"):
        file_storage.seek(0)
    data = file_storage.read()
    if hasattr(file_storage, "seek"):
        file_storage.seek(0)
    return data


def pdf_to_text(file_storage):
    data = _read_file_bytes(file_storage)
    reader = PdfReader(io.BytesIO(data))
    text = ""
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text += page_text + "\n"
    return text.strip()


def docx_to_text(file_storage):
    data = _read_file_bytes(file_storage)
    doc = Document(io.BytesIO(data))
    parts = []
    for paragraph in doc.paragraphs:
        if paragraph.text.strip():
            parts.append(paragraph.text)
    return "\n".join(parts).strip()


def excel_to_text(file_storage, filename=""):
    ext = os.path.splitext(filename)[1].lower()
    data = _read_file_bytes(file_storage)

    if ext == ".csv":
        text = data.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        return "\n".join(", ".join(row) for row in reader).strip()

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in workbook.worksheets:
        parts.append(f"[{sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None and str(cell).strip()]
            if cells:
                parts.append(" | ".join(cells))
    workbook.close()
    return "\n".join(parts).strip()


def text_file_to_text(file_storage):
    return _read_file_bytes(file_storage).decode("utf-8", errors="ignore").strip()


def image_to_data_url(file_storage):
    mime_type = file_storage.mimetype or "image/jpeg"
    raw = _read_file_bytes(file_storage)
    encoded = base64.b64encode(raw).decode("utf-8")
    return f"data:{mime_type};base64,{encoded}"


def extract_file_text(file_storage, filename, mimetype=""):
    category = get_file_category(filename, mimetype)
    if category == "pdf":
        return pdf_to_text(file_storage), category
    if category == "word":
        return docx_to_text(file_storage), category
    if category == "excel":
        return excel_to_text(file_storage, filename), category
    if category == "text":
        return text_file_to_text(file_storage), category
    if category == "image":
        return None, category
    raise ValueError(f"Desteklenmeyen dosya türü: {filename}")


def build_file_meta(filename, category, preview=None):
    return {
        "name": filename,
        "type": category,
        "icon": file_icon_for(category),
        "preview": preview,
    }


def parse_uploaded_attachment(uploaded_file, max_size=10 * 1024 * 1024):
    if not uploaded_file or not uploaded_file.filename:
        return None

    data = _read_file_bytes(uploaded_file)
    if len(data) > max_size:
        raise ValueError("Dosya boyutu 10 MB sınırını aşıyor.")

    return {
        "filename": uploaded_file.filename,
        "mimetype": uploaded_file.mimetype or "application/octet-stream",
        "data": data,
    }
